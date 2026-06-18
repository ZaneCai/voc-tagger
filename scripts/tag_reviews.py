#!/usr/bin/env python3
"""
Amazon VOC Tagger — 核心打标脚本
使用 CRS claude-sonnet-4-6 + tool call + enum 约束

使用前：
1. 修改下方 CONFIG 区块（路径、API Key、代理）
2. 修改 TAXONOMY 区块（你的三级标签体系）
3. 修改 build_enum() 中的 L2_MAP / L3_MAP（与 TAXONOMY 一一对应）
4. python3 -u tag_reviews.py
"""

import json, time, urllib.request, openpyxl
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

# ── CONFIG ────────────────────────────────────────────────────────────
CRS_BASE     = 'https://your-openai-compatible-endpoint/api'  # 替换为你的 OpenAI-compatible API 地址
CRS_KEY      = ''          # 填入你的 API Key，或从环境变量/secrets 读取
PROXY        = ''          # 选填：如需代理访问填自己的端口，如 http://127.0.0.1:<port>
MODEL        = 'claude-sonnet-4-6'
BATCH        = 10          # 每批条数（建议 8-12）
INPUT_FILE   = ''          # 输入 Excel 路径
OUTPUT_FILE  = ''          # 输出 Excel 路径
CHECKPOINT   = 'tag_checkpoint.json'

# 输入字段名（按实际 Excel 表头修改）
COL_ID       = 'id'
COL_STAR     = 'star'
COL_MARKET   = 'market'
COL_TITLE    = 'title_zh'
COL_TEXT     = 'text_zh'

# ── TAXONOMY ──────────────────────────────────────────────────────────
# 按实际标签体系修改，格式：
# 一级 > 二级(三级|三级) | 二级(三级)
TAXONOMY = """
0.正向反馈: 0.1速度/性能满意>0.1.1有线速度满意,0.1.2无线速度满意,0.1.3游戏低延迟
1.稳定性: 1.1WiFi随机断线>1.1.1全频段掉线重启恢复,1.1.2特定设备反复掉线
"""
# TODO: 替换为完整的标签体系

# Few-shot 示例（3条，帮助模型理解 l1/l2/l3 格式）
FEWSHOT = """
示例1：
输入：[ID:101] ★1 US 标题：风扇很吵 正文：风扇全速转，影响睡眠，退货
输出tags：[{"l1":"4.噪音与发热","l2":"4.1风扇噪音大","l3":"4.1.2影响睡眠/生活"}]

示例2：
输入：[ID:102] ★5 US 标题：速度飞快 正文：无线速度超过1Gbps，设置很简单
输出tags：[{"l1":"0.正向反馈","l2":"0.1速度/性能满意","l3":"0.1.2无线速度满意"},{"l1":"0.正向反馈","l2":"0.4设置简单","l3":"0.4.1App引导顺畅"}]

示例3：
输入：[ID:103] ★1 US 标题：频繁断线 正文：WiFi每30分钟断一次，重启才恢复
输出tags：[{"l1":"1.稳定性","l2":"1.1WiFi随机断线","l3":"1.1.1全频段掉线重启恢复"}]
"""

# ── ENUM 定义（与 TAXONOMY 一一对应）─────────────────────────────────
# 修改为完整的合法值列表
def build_enums():
    L1_ENUM = [
        "0.正向反馈", "1.稳定性",
        # TODO: 补充其他一级标签
    ]

    L2_MAP = {
        "0.正向反馈": ["0.1速度/性能满意", "0.4设置简单"],
        "1.稳定性":   ["1.1WiFi随机断线"],
        # TODO: 补充其他
    }

    L3_MAP = {
        "0.1速度/性能满意": ["0.1.1有线速度满意", "0.1.2无线速度满意", "0.1.3游戏低延迟"],
        "0.4设置简单":      ["0.4.1App引导顺畅"],
        "1.1WiFi随机断线":  ["1.1.1全频段掉线重启恢复", "1.1.2特定设备反复掉线"],
        # TODO: 补充其他
    }

    ALL_L2 = [v for vals in L2_MAP.values() for v in vals]
    ALL_L3 = [v for vals in L3_MAP.values() for v in vals]
    return L1_ENUM, ALL_L2, ALL_L3


# ── TOOL DEFINITION ───────────────────────────────────────────────────
def build_tool(L1_ENUM, ALL_L2, ALL_L3):
    return [{
        'type': 'function',
        'function': {
            'name': 'submit_tags',
            'description': '提交所有评论的VOC标签结果',
            'parameters': {
                'type': 'object',
                'properties': {
                    'results': {
                        'type': 'array',
                        'items': {
                            'type': 'object',
                            'properties': {
                                'id':   {'type': 'integer'},
                                'tags': {
                                    'type': 'array',
                                    'items': {
                                        'type': 'object',
                                        'properties': {
                                            'l1': {'type': 'string', 'enum': L1_ENUM},
                                            'l2': {'type': 'string', 'enum': ALL_L2},
                                            'l3': {'type': 'string', 'enum': ALL_L3},
                                        },
                                        'required': ['l1', 'l2', 'l3']
                                    }
                                }
                            },
                            'required': ['id', 'tags']
                        }
                    }
                },
                'required': ['results']
            }
        }
    }]


# ── API CALL ──────────────────────────────────────────────────────────
def call_crs(messages, tools, system_prompt):
    proxy_handler = urllib.request.ProxyHandler({'https': PROXY, 'http': PROXY})
    opener = urllib.request.build_opener(proxy_handler)

    payload = json.dumps({
        'model': MODEL,
        'messages': [{'role': 'system', 'content': system_prompt}] + messages,
        'tools': tools,
        'tool_choice': {'type': 'function', 'function': {'name': 'submit_tags'}},
        'max_tokens': 2000,
        'temperature': 0
    }).encode()

    req = urllib.request.Request(
        f'{CRS_BASE}/v1/chat/completions', data=payload,
        headers={'Authorization': f'Bearer {CRS_KEY}', 'Content-Type': 'application/json',
                 'User-Agent': 'Mozilla/5.0'}
    )
    with opener.open(req, timeout=120) as r:
        resp = json.loads(r.read())

    msg = resp['choices'][0]['message']
    if 'tool_calls' in msg:
        return json.loads(msg['tool_calls'][0]['function']['arguments']).get('results', [])
    return []


# ── MAIN ──────────────────────────────────────────────────────────────
def load_reviews():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    reviews = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        reviews.append({
            'raw': list(row),
            'id': row[idx[COL_ID]],
            'star': row[idx[COL_STAR]],
            'market': row[idx.get(COL_MARKET, -1)] if COL_MARKET in idx else '',
            'title': row[idx[COL_TITLE]] or '',
            'text': row[idx[COL_TEXT]] or '',
        })
    return headers, reviews


def load_checkpoint():
    try:
        with open(CHECKPOINT) as f:
            return json.load(f)
    except:
        return []


def save_checkpoint(all_tags):
    with open(CHECKPOINT, 'w') as f:
        json.dump(all_tags, f, ensure_ascii=False)


def build_output(headers, reviews, all_tags):
    tag_map = {str(t['id']): t.get('tags', []) for t in all_tags}
    rows_out = []
    for r in reviews:
        tags = tag_map.get(str(r['id']), [])
        if not tags:
            rows_out.append(r['raw'] + ['', '', ''])
        else:
            for tag in tags:
                rows_out.append(r['raw'] + [tag.get('l1',''), tag.get('l2',''), tag.get('l3','')])
    return rows_out


def save_excel(headers, rows_out):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'tagged'
    new_headers = headers + ['l1', 'l2', 'l3']
    ws.append(new_headers)
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(color='FFFFFF', bold=True)
    for row in rows_out:
        ws.append(row)
    wb.save(OUTPUT_FILE)
    print(f'\n✅ {OUTPUT_FILE}，共 {len(rows_out)} 行', flush=True)


def main():
    assert CRS_KEY,     '请配置 CRS_KEY'
    assert INPUT_FILE,  '请配置 INPUT_FILE'
    assert OUTPUT_FILE, '请配置 OUTPUT_FILE'

    L1_ENUM, ALL_L2, ALL_L3 = build_enums()
    tools = build_tool(L1_ENUM, ALL_L2, ALL_L3)

    system_prompt = f"""你是VOC分析师，为产品评论打标签。

标签体系（必须严格使用，不得自创）：
{TAXONOMY}

示例：
{FEWSHOT}

调用 submit_tags 函数提交结果，标签必须来自体系。"""

    print('📖 读取数据...', flush=True)
    headers, reviews = load_reviews()
    print(f'共 {len(reviews)} 条评论', flush=True)

    all_tags = load_checkpoint()
    tagged_ids = {str(t['id']) for t in all_tags}
    if all_tags:
        print(f'已有 checkpoint：{len(all_tags)} 条', flush=True)

    total_batches = (len(reviews) + BATCH - 1) // BATCH

    for i in range(0, len(reviews), BATCH):
        batch = reviews[i:i+BATCH]
        batch_num = i // BATCH + 1

        if all(str(r['id']) in tagged_ids for r in batch):
            print(f'  跳过批次 {batch_num}（已完成）', flush=True)
            continue

        print(f'\n🏷️  批次 {batch_num}/{total_batches}（第 {i+1}~{min(i+BATCH, len(reviews))} 条）...', flush=True)

        lines = []
        for r in batch:
            lines.append(f'[ID:{r["id"]}] ★{r["star"]} {r["market"]} 标题：{r["title"]} 正文：{r["text"][:500]}')

        user_msg = '请为以下评论打标签：\n\n' + '\n\n'.join(lines)

        for attempt in range(3):
            try:
                result = call_crs([{'role': 'user', 'content': user_msg}], tools, system_prompt)
                all_tags.extend(result)
                tagged_ids.update(str(t['id']) for t in result)
                save_checkpoint(all_tags)
                print(f'  ✓ 返回 {len(result)} 条，累计 {len(all_tags)} 条', flush=True)
                break
            except Exception as e:
                print(f'  Attempt {attempt+1} failed: {e}', flush=True)
                if attempt < 2:
                    time.sleep(3)

        if batch_num < total_batches:
            time.sleep(2)

    print(f'\n📊 共 {len(all_tags)} 条标签结果，生成 Excel...', flush=True)
    rows_out = build_output(headers, reviews, all_tags)
    save_excel(headers, rows_out)


if __name__ == '__main__':
    main()
